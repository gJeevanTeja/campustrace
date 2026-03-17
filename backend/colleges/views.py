from rest_framework import viewsets, permissions, status
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Count, Q, Avg, F
from django.db.models.functions import ExtractHour
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
import csv
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from .models import College, Block, Category, CampusLocation
from rest_framework import generics
from .serializers import CollegeSerializer, BlockSerializer, CategorySerializer, CampusLocationSerializer, ItemReportSerializer
from users.permissions import IsSuperAdmin, IsCollegeAdmin, IsAdminOrModerator
from items.models import Item

class ItemReportListView(generics.ListAPIView):
    serializer_class = ItemReportSerializer
    permission_classes = [IsAdminOrModerator]

    def get_queryset(self):
        view = BaseExportView()
        return view.get_queryset_filtered(self.request)

class BaseExportView(APIView):
    permission_classes = [IsAdminOrModerator]

    def get_queryset_filtered(self, request):
        user = request.user
        college = getattr(user, 'college', None)
        
        item_filter = Q()
        if user.role != 'super_admin':
            if not college:
                return Item.objects.none()
            item_filter &= Q(college=college)

        # Apply Filters from query params
        query = request.query_params.get('search', '').strip()
        category_id = request.query_params.get('category', '').strip()
        item_type = request.query_params.get('type', '').strip()
        status = request.query_params.get('status', '').strip()
        college_id = request.query_params.get('collegeId', '').strip()
        start_date = request.query_params.get('startDate', '').strip()
        end_date = request.query_params.get('endDate', '').strip()

        if query:
            item_filter &= (Q(title__icontains=query) | Q(description__icontains=query) | Q(reference_number__icontains=query))
        
        if category_id:
            item_filter &= Q(category_new_id=category_id)
            
        if item_type:
            item_filter &= Q(type=item_type)
            
        if status:
            item_filter &= Q(status=status)

        if user.role == 'super_admin' and college_id:
            item_filter &= Q(college_id=college_id)

        if start_date:
            item_filter &= Q(created_at__gte=start_date)
        if end_date:
            item_filter &= Q(created_at__lte=end_date)
            
        return Item.objects.filter(item_filter).select_related('user', 'category_new', 'block', 'college').order_by('-created_at')

    def get_report_data(self, request):
        items = self.get_queryset_filtered(request)
        if items is None:
            return None
            
        data = []
        for item in items:
            claim = item.claim_sessions.first()
            claim_status = claim.status.title() if claim else "No Claim"
            
            data.append({
                'item_name': item.title,
                'category': item.category_new.name if item.category_new else item.get_category_display(),
                'reported_by': item.user.name or item.user.username,
                'type': item.get_type_display(),
                'status': item.get_status_display(),
                'location': item.block.name if item.block else item.get_location_display(),
                'date_reported': item.created_at.strftime("%Y-%m-%d %H:%M"),
                'claim_status': claim_status
            })
        return data

def create_default_college_data(college):
    """Creates default categories and locations for a new college."""
    # Categories
    default_categories = ["ID Card", "Bags", "Helmet", "Keys", "Mobile Phones", "Other"]
    for name in default_categories:
        Category.objects.get_or_create(name=name, college=college)
        
    # Locations / Blocks
    # Default to Malla Reddy University center (approx 17.5583, 78.4300)
    default_locations = ["Canteen", "Parking", "Gym Area", "Ground", "Main Gate", "Library", "Auditorium", "Other"]
    for name in default_locations:
        Block.objects.get_or_create(name=name, college=college)
        CampusLocation.objects.get_or_create(
            name=name, 
            college=college,
            latitude=17.5583,
            longitude=78.4300
        )

class CollegeViewSet(viewsets.ModelViewSet):
    queryset = College.objects.all()
    serializer_class = CollegeSerializer
    
    def get_permissions(self):
        if self.action == 'list':
            return [permissions.AllowAny()]
        return [IsSuperAdmin()]

    def perform_create(self, serializer):
        college = serializer.save()
        create_default_college_data(college)

class BlockViewSet(viewsets.ModelViewSet):
    queryset = Block.objects.all()
    serializer_class = BlockSerializer

    def get_permissions(self):
        if self.action == 'list':
            return [permissions.AllowAny()]
        return [IsAdminOrModerator()]

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated and user.role == 'super_admin':
            return Block.objects.all()
        if user.is_authenticated and getattr(user, 'college', None):
            return Block.objects.filter(college=user.college, is_active=True)
        return Block.objects.filter(is_active=True)

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == 'super_admin':
            college_id = self.request.data.get('college')
            if college_id:
                serializer.save()
            else:
                from .models import College
                college = College.objects.first()
                serializer.save(college=college)
        else:
            serializer.save(college=user.college)

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return Response({
            "success": True,
            "message": "Block created successfully",
            "data": response.data
        }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        return Response({
            "success": True,
            "message": "Block updated successfully",
            "data": response.data
        }, status=status.HTTP_200_OK)

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer

    def get_permissions(self):
        if self.action == 'list':
            return [permissions.AllowAny()]
        return [IsAdminOrModerator()]

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated and user.role == 'super_admin':
            return Category.objects.all()
        if user.is_authenticated and getattr(user, 'college', None):
            return Category.objects.filter(college=user.college, active=True)
        # Fallback
        return Category.objects.filter(active=True)

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == 'super_admin':
            # For super admin, use college from data or fall back to first college
            college_id = self.request.data.get('college')
            if college_id:
                serializer.save() # DRF will use 'college' from validated_data
            else:
                from .models import College
                college = College.objects.first()
                serializer.save(college=college)
        else:
            serializer.save(college=user.college)

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return Response({
            "success": True,
            "message": "Category created successfully",
            "data": response.data
        }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        return Response({
            "success": True,
            "message": "Category updated successfully",
            "data": response.data
        }, status=status.HTTP_200_OK)

from rest_framework import generics

class CampusLocationListView(generics.ListAPIView):
    serializer_class = CampusLocationSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated and user.role == 'super_admin':
            return CampusLocation.objects.all()
        if user.is_authenticated and getattr(user, 'college', None):
            return CampusLocation.objects.filter(college=user.college)
        # Fallback to all if no college assigned
        return CampusLocation.objects.all()

class AdminAnalyticsView(APIView):
    permission_classes = [IsAdminOrModerator]

    def get(self, request):
        user = request.user
        college = user.college
        
        # Base filter
        item_filter = Q()
        if user.role != 'super_admin':
            if not college:
                return Response({'error': 'User not linked to any college'}, status=status.HTTP_400_BAD_REQUEST)
            item_filter &= Q(college=college)
            
        items = Item.objects.filter(item_filter)
        total_items = items.count()
        total_lost = items.filter(type='lost').count()
        total_found = items.filter(type='found').count()
        
        # Resolution Rate (Returned / Total)
        resolved_count = items.filter(status='returned').count()
        res_rate = (resolved_count / total_items * 100) if total_items > 0 else 0
        
        # Category stats for BarChart
        category_stats = list(items.values(name=F('category_new__name')).annotate(count=Count('id')).order_by('-count')[:8])
        
        # Most Lost Category
        most_lost = items.filter(type='lost').values('category_new__name').annotate(count=Count('id')).order_by('-count').first()
        most_lost_name = most_lost['category_new__name'] if most_lost and most_lost['category_new__name'] else "N/A"
        
        # Peak Hour (Number)
        peak_hour_data = items.annotate(hour=ExtractHour('created_at')).values('hour').annotate(count=Count('id')).order_by('-count').first()
        peak_hour = peak_hour_data['hour'] if peak_hour_data else 0

        # Active Users (last 7 days per user request)
        from users.models import User
        active_users = User.objects.filter(last_active__gte=timezone.now() - timedelta(days=7))
        if user.role != 'super_admin':
            active_users = active_users.filter(college=college)
        active_count = active_users.count()
        
        # Avg Return Time
        resolved_items = items.filter(status='returned').annotate(
            duration=F('updated_at') - F('created_at')
        ).aggregate(avg_time=Avg('duration'))
        
        avg_display = "N/A"
        if resolved_items['avg_time']:
            td = resolved_items['avg_time']
            days = td.days
            hours = td.seconds // 3600
            if days > 0:
                avg_display = f"{days}d {hours}h"
            else:
                avg_display = f"{hours}h"

        # Reporting Activity (Last 7 days)
        activity_data = []
        for i in range(7):
            date = timezone.now().date() - timedelta(days=i)
            count = items.filter(created_at__date=date).count()
            activity_data.append({
                "date": date.strftime("%b %d"),
                "count": count
            })
        activity_data.reverse()

        return Response({
            "totalReports": total_items,
            "resolutionRate": round(res_rate, 1),
            "activeUsers": active_count,
            "avgReturnTime": avg_display,
            "reportsByDay": activity_data,
            "lostVsFound": {
                "lost": total_lost,
                "found": total_found
            },
            "peakHour": peak_hour,
            "topCategory": most_lost_name,
            "categoryStats": category_stats
        })


class ExportCSVView(BaseExportView):
    def get(self, request):
        report_data = self.get_report_data(request)
        if report_data is None:
            return Response({'error': 'User not linked to any college'}, status=status.HTTP_400_BAD_REQUEST)

        response = HttpResponse(content_type='text/csv')
        filename = f"CampusTrace_Report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(['Item Name', 'Category', 'Type', 'Reported By', 'Status', 'Location', 'Date Reported', 'Claim Status'])

        for row in report_data:
            writer.writerow([
                row['item_name'],
                row['category'],
                row['type'],
                row['reported_by'],
                row['status'],
                row['location'],
                row['date_reported'],
                row['claim_status']
            ])

        return response


class ExportExcelView(BaseExportView):
    def get(self, request):
        report_data = self.get_report_data(request)
        if report_data is None:
            return Response({'error': 'User not linked to any college'}, status=status.HTTP_400_BAD_REQUEST)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CampusTrace Report"

        # Headers
        headers = ['Item Name', 'Category', 'Type', 'Reported By', 'Status', 'Location', 'Date Reported', 'Claim Status']
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data
        for row in report_data:
            ws.append([
                row['item_name'],
                row['category'],
                row['type'],
                row['reported_by'],
                row['status'],
                row['location'],
                row['date_reported'],
                row['claim_status']
            ])

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"CampusTrace_Report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ExportPDFView(BaseExportView):
    def get(self, request):
        report_data = self.get_report_data(request)
        if report_data is None:
            return Response({'error': 'User not linked to any college'}, status=status.HTTP_400_BAD_REQUEST)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []

        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        title_style.alignment = 1 # Center
        
        # Add Header
        elements.append(Paragraph("CampusTrace Admin Report", title_style))
        elements.append(Spacer(1, 12))
        
        # Add Statistics
        items_count = len(report_data)
        lost_count = sum(1 for r in report_data if r['type'] == 'Lost')
        found_count = sum(1 for r in report_data if r['type'] == 'Found')
        returned_count = sum(1 for r in report_data if r['status'] == 'Returned')

        stats_text = f"Total Items: {items_count} | Lost: {lost_count} | Found: {found_count} | Returned: {returned_count}"
        elements.append(Paragraph(stats_text, styles['Normal']))
        elements.append(Spacer(1, 24))

        # Table Data
        data = [['Item', 'Category', 'Type', 'Status', 'Location', 'Date']]
        for row in report_data:
            # Shorten date for PDF table space
            short_date = row['date_reported'].split(' ')[0]
            data.append([
                Paragraph(row['item_name'], styles['Normal']),
                row['category'],
                row['type'],
                row['status'],
                row['location'],
                short_date
            ])

        # Create Table
        table = Table(data, colWidths=[100, 80, 50, 70, 90, 70])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        filename = f"CampusTrace_Report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
